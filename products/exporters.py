# -- coding: utf-8 --
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.writer.excel import save_virtual_workbook
from numpy import array
import barcode
from barcode.writer import ImageWriter


class Exporter(object):

    def export(self, model):
        raise NotImplementedError()


class BoxExporter(Exporter):
    metadata_font = Font(
        size=20,
        bold=True
    )

    table_header_font = Font(
        bold=True
    )

    content_type = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'

    def _write_customer_title(self, work_sheet, customer):
        work_sheet.append((
            'Customer', str(customer),
        ))

        row = self._get_last_row(work_sheet)
        row.font = self.metadata_font

        if customer.invoice:
            work_sheet.append((
                'Invoice', customer.invoice,
            ))

            row = self._get_last_row(work_sheet)
            row.font = self.metadata_font
        else:
            self._write_empty_line(work_sheet)


    def _write_box_title(self, work_sheet, box):
        work_sheet.append((
            'Box', str(box),
        ))

        row = self._get_last_row(work_sheet)
        row.font = self.metadata_font

    def _write_customer_total_weight(self, work_sheet, customer):
        if customer.total_weight:
            work_sheet.append((
                'Total weight', customer.total_weight
            ))

            row = self._get_last_row(work_sheet)
            row.font = self.metadata_font

    def _write_product_table_header(self, work_sheet):
        work_sheet.append((
            'Barcode',
            'Name',
            'Order',
            'Quantity',
        ))

        row = self._get_last_row(work_sheet)
        row.font = self.table_header_font

    def _write_empty_line(self, work_sheet):
        work_sheet.append([])

    def _get_last_row(self, work_sheet):
        rows = work_sheet.row_dimensions
        return rows[max(rows.keys())]

    def _write_product_table_row(self, work_book, work_sheet, product, line):

        barcode_code = product.barcode
        EAN = barcode.get_barcode_class('ean13')
        ean = EAN(barcode_code, writer=ImageWriter())
        file_name = ean.save(product.name)
        size = openpyxl.drawing.Image(file_name).image.size * array([0.5, 0.5])
        img = openpyxl.drawing.Image(file_name, size=size)
        line_number = 'A%s' %line
        img.anchor(work_book.worksheets[0].cell(line_number))


        work_sheet.append((
            product.name,
            product.order,
            product.quantity,
        ))
        work_sheet.append((work_sheet.add_image(img),)),

    def export(self, box):
        work_book = Workbook()

        work_sheet = work_book.active
        work_sheet.title = 'Box'

        self._write_customer_title(work_sheet, box.parent)
        self._write_box_title(work_sheet, box)
        self._write_empty_line(work_sheet)
        self._write_product_table_header(work_sheet)

        for product in box.productpurchase_set.all():
            self._write_product_table_row(work_sheet, product)

        return save_virtual_workbook(work_book)


class CustomerExporter(BoxExporter):

    def export(self, customer):
        work_book = Workbook()

        work_sheet = work_book.active
        work_sheet.column_dimensions['A'].width = 35
        work_sheet.column_dimensions['B'].width = 60
        work_sheet.column_dimensions['C'].width = 25
        work_sheet.column_dimensions['D'].width = 13
        work_sheet.title = 'Boxs'

        self._write_customer_title(work_sheet, customer)
        self._write_customer_total_weight(work_sheet, customer)
        self._write_empty_line(work_sheet)
        work_sheet.row_dimensions[1].height = 30
        work_sheet.row_dimensions[2].height = 30
        work_sheet.row_dimensions[3].height = 30

        line = 5

        for box in customer.box_set.all():
            self._write_box_title(work_sheet, box)
            self._write_product_table_header(work_sheet)
            for product in box.productpurchase_set.all():
                self._write_product_table_row(work_book, work_sheet, product, line)
                work_sheet.row_dimensions[line].height = 100
                line  = line+1
            self._write_empty_line(work_sheet)
            line  = line+2


        file_name = '%s.xlsx' % str(customer)
        work_book.save(file_name)
        return save_virtual_workbook(work_book)


from .models import Box, Customer

__exporters_instances = {
    Box: BoxExporter(),
    Customer: CustomerExporter(),
}

def get_exporter(model):
    return __exporters_instances[model]
